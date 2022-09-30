#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2022/8/30 
# @Author  : henry
# @File    : exer_4.py
import openpyxl

""" -------------homework---------------"""

# 1、通过上述上课所学内置函数和推导式的语法，读取附件中excel文件，转换为如下的格式：
# res1 = [ {'case_id': 1, 'case_title': '用例1', 'url': 'www.baidu.com', 'data': '001', 'excepted': 'ok'},
#  {'case_id': 4, 'case_title': '用例4', 'url': 'www.baidu.com', 'data': '002', 'excepted': 'ok'},
#  {'case_id': 2, 'case_title': '用例2', 'url': 'www.baidu.com', 'data': '002', 'excepted': 'ok'},
#  {'case_id': 3, 'case_title': '用例3', 'url': 'www.baidu.com', 'data': '002', 'excepted': 'ok'},
#  {'case_id': 5, 'case_title': '用例5', 'url': 'www.baidu.com', 'data': '002', 'excepted': 'ok'} ]
# 2、对第一题读取出来的数据，按照case_id字段进行排序。
# 3、将读取出来的数据中的method字段值 统一修改为GET（不需要修改excel，只对读出来的数据进行修改）
# 1.方式1
book = openpyxl.load_workbook('data.xlsx')
sheet = book.active
lines = sheet.values  # 按行获取所有单元格的值
title = next(lines)
res1 = []
for each in lines:
    res1.append(dict(zip(title, each)))
print(res1)

# 方式2
rows = openpyxl.load_workbook('data.xlsx')['login'].rows  # 行数据的迭代器
title = [i.value for i in next(rows)]  # 迭代一次读标题
res2 = [dict(zip(title, [i.value for i in item])) for item in rows]  # 余下的行数据(单元格)继续迭代
print(res2)
#  方式3
rows = list(openpyxl.load_workbook('data.xlsx')['login'].rows)  # 行数据的迭代器列表化
res3 = [dict(zip([i.value for i in rows[0]], [i.value for i in item])) for item in rows[1:]]
print(res3)

# 方式4
rows = list(openpyxl.load_workbook('data.xlsx')['login'].rows)  # 行数据的迭代器列表化
#  使用map
res4 = [dict(zip(map(lambda x: x.value, rows[0]), map(lambda x: x.value, item))) for item in rows[1:]]
print(res4)

# 2.
# 方式1
res1.sort(key=lambda x: x['case_id'])
print(res1)


# 3.方式1 常规
# for it in res1:
#     it['method'] = 'GET'
# print(res1)

# 方式2 map方法，返回一个map对象，被map的对象数据不会被修改
# lambda 中不能使用‘=’赋值，使用exec函数去执行, exec返回None 与传入数据进行or操作，得修改后的数据
res2 = map(lambda x: exec("x['method'] = 'GET'") or x, res1)
# print(list(res2))

# 方式3 map方法 update方法
res3 = map(lambda x: x.update(method='GET') or x, res1)
# print(list(res3))

# 方式4 update方法 新列表
result4 = [item.update({'method': 'GET'}) or item for item in res1]
print(result4)


def work2_sort(func):

    def wrapper():
        data = func()
        data.sort(key=lambda x: x['case_id'])
        return data

    return wrapper


def work3_update(func):

    def wrapper():
        data = func()
        for every in data:
            every['method'] = 'GET'
        return data
    return wrapper


# 装饰器语法实现
@work3_update
@work2_sort
def work1_read():
    # 读取excel数据函数
    workbook = openpyxl.load_workbook('data.xlsx')
    row = workbook.active.rows
    titles = next(row)
    return [dict(zip([i.value for i in titles], [j.value for j in item])) for item in row]


print(work1_read())
